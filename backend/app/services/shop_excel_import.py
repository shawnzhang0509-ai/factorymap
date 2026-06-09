"""Parse .xlsx bulk import files for factory rows (admin / ad manager)."""
from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook

# (canonical_field, header labels — matched case-insensitively after strip)
_HEADER_GROUPS: list[tuple[str, tuple[str, ...]]] = [
    ("name", ("name", "名称", "工厂名", "店名", "店铺名称", "企业名称")),
    ("address", ("address", "地址", "详细地址")),
    ("phone", ("phone", "电话", "手机", "联系电话")),
    ("lat", ("lat", "latitude", "纬度")),
    ("lng", ("lng", "lon", "longitude", "经度")),
    ("province", ("省份", "省")),
    ("city", ("城市", "市")),
    ("district", ("区县", "区", "县")),
    ("badge_text", ("badge_text", "tags", "标签", "徽章")),
    (
        "new_girls_last_15_days",
        ("new_girls_last_15_days", "new_badge", "新店", "显示新店"),
    ),
    ("about_me", ("about_me", "简介", "about", "描述", "description")),
    ("additional_price", ("additional_price", "附加费用", "加价说明")),
    ("filter_city", ("filter_city", "区域", "筛选城市", "产业带")),
    ("min_spend", ("min_spend", "moq", "起订量", "最低消费")),
    ("main_product", ("main_product", "主营产品", "产品", "品类")),
    ("industry", ("行业分类", "行业")),
    ("photo_url", ("照片url", "照片 url", "图片url", "图片地址")),
]

# Extra columns folded into about_me (factory profile / description)
_DESCRIPTION_LINES: list[tuple[str, tuple[str, ...]]] = [
    ("统一社会信用代码", ("统一社会信用代码", "信用代码", "社会信用代码")),
    ("注册资本", ("注册资本",)),
    ("企业状态", ("企业状态", "经营状态")),
    ("数据来源", ("数据来源", "来源")),
    ("行业分类", ("行业分类",)),
    ("法人/负责人", ("法人/负责人", "法人", "负责人")),
    ("网址", ("网址", "网站", "website", "官网")),
    ("成立日期", ("成立日期", "注册日期")),
    ("照片URL", ("照片url", "照片 url", "图片url")),
]

_HEADER_TO_FIELD: dict[str, str] = {}
for field, labels in _HEADER_GROUPS:
    for label in labels:
        key = label.strip().lower()
        if key and key not in _HEADER_TO_FIELD:
            _HEADER_TO_FIELD[key] = field

_DESC_HEADER_TO_LABEL: dict[str, str] = {}
for label_zh, aliases in _DESCRIPTION_LINES:
    for alias in aliases:
        key = alias.strip().lower()
        if key and key not in _DESC_HEADER_TO_LABEL:
            _DESC_HEADER_TO_LABEL[key] = label_zh

MAX_IMPORT_ROWS = 500

PEARL_RIVER_DELTA = "Pearl River Delta"
YANGTZE_DELTA = "Yangtze River Delta"
BOHAI_RIM = "Bohai Economic Rim"
CENTRAL_WEST = "Central & Western China"


def _norm_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\ufeff", "").strip().lower()


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        if isinstance(value, float) and value == int(value):
            return str(int(value))
        return str(value).strip()
    s = str(value).strip()
    return s if s else None


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", "."))
    except (TypeError, ValueError):
        return None


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value != 0
    if isinstance(value, float):
        return value != 0.0
    s = (str(value).strip().lower() if value is not None else "")
    return s in ("1", "true", "yes", "y", "是", "对")


def _normalize_phone(raw: str | None) -> str | None:
    if not raw:
        return None
    s = str(raw).strip()
    if not s:
        return None
    # Keep first segment if multiple numbers in one cell
    for sep in (";", "；", "/", "、", "\n"):
        if sep in s:
            s = s.split(sep)[0].strip()
    return s[:20] if len(s) > 20 else s


def _compose_address(
    detailed: str | None,
    province: str | None,
    city: str | None,
    district: str | None,
) -> str:
    parts: list[str] = []
    for p in (province, city, district):
        if p and p not in parts:
            parts.append(p)
    base = "".join(parts)
    detail = (detailed or "").strip()
    if detail:
        if base and detail not in base and not detail.startswith(base):
            return f"{base}{detail}"
        return detail if detail else base
    return base


def _infer_filter_city(
    province: str | None,
    city: str | None,
    district: str | None,
    address: str | None,
    data_source: str | None,
    explicit: str | None,
) -> str | None:
    if explicit and explicit.strip():
        return explicit.strip()
    hay = "".join(
        x or ""
        for x in (province, city, district, address, data_source)
    )
    if not hay:
        return None
    pearl = ("广东", "佛山", "顺德", "深圳", "东莞", "广州", "珠海", "惠州", "珠三角", "产业带_佛山")
    yangtze = ("江苏", "浙江", "上海", "苏州", "无锡", "南京", "杭州", "长三角")
    bohai = ("山东", "河北", "天津", "北京", "环渤海")
    west = ("四川", "重庆", "湖北", "湖南", "陕西", "成都", "武汉", "中西部")
    if any(m in hay for m in pearl):
        return PEARL_RIVER_DELTA
    if any(m in hay for m in yangtze):
        return YANGTZE_DELTA
    if any(m in hay for m in bohai):
        return BOHAI_RIM
    if any(m in hay for m in west):
        return CENTRAL_WEST
    return None


def _build_description_block(
    header_to_col: dict[str, int],
    row: tuple[Any, ...],
    header_row: tuple[Any, ...],
) -> str:
    lines: list[str] = []
    seen_labels: set[str] = set()

    for idx, cell in enumerate(header_row):
        h = _norm_header(cell)
        if not h:
            continue
        label = _DESC_HEADER_TO_LABEL.get(h)
        if not label or label in seen_labels:
            continue
        if idx >= len(row):
            continue
        val = _cell_str(row[idx])
        if not val:
            continue
        seen_labels.add(label)
        lines.append(f"{label}: {val}")

    return "\n".join(lines)


def _row_to_payload(
    header_to_col: dict[str, int],
    row: tuple[Any, ...],
    header_row: tuple[Any, ...],
) -> dict[str, Any]:
    def get(field: str) -> Any:
        idx = header_to_col.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    name = _cell_str(get("name"))
    province = _cell_str(get("province"))
    city = _cell_str(get("city"))
    district = _cell_str(get("district"))
    address = _compose_address(
        _cell_str(get("address")),
        province,
        city,
        district,
    )
    phone = _normalize_phone(_cell_str(get("phone")))
    lat = _parse_float(get("lat"))
    lng = _parse_float(get("lng"))
    badge = _cell_str(get("badge_text"))
    about = _cell_str(get("about_me")) or ""
    add_price = _cell_str(get("additional_price")) or ""
    filter_city = _cell_str(get("filter_city")) or ""
    min_raw = get("min_spend")
    min_str = _cell_str(min_raw) if min_raw is not None else None
    main_product = _cell_str(get("main_product")) or _cell_str(get("industry")) or ""

    raw_new = get("new_girls_last_15_days")
    new_girls = _parse_bool(raw_new) if raw_new not in (None, "") else False

    desc_block = _build_description_block(header_to_col, row, header_row)
    industry = _cell_str(get("industry"))
    if industry:
        industry_line = f"行业分类: {industry}"
        if industry_line not in desc_block:
            desc_block = f"{desc_block}\n{industry_line}".strip() if desc_block else industry_line
    if desc_block:
        about = f"{about}\n\n{desc_block}".strip() if about else desc_block

    data_source = None
    for idx, cell in enumerate(header_row):
        h = _norm_header(cell)
        if h in ("数据来源", "来源") and idx < len(row):
            data_source = _cell_str(row[idx])
            break

    inferred_zone = _infer_filter_city(
        province, city, district, address, data_source, filter_city
    )

    data: dict[str, Any] = {}
    if name is not None:
        data["name"] = name[:100]
    if address:
        data["address"] = address[:200]
    if phone is not None:
        data["phone"] = phone
    if lat is not None:
        data["lat"] = lat
    if lng is not None:
        data["lng"] = lng
    if badge is not None:
        data["badge_text"] = badge
    data["new_girls_last_15_days"] = new_girls
    if about:
        data["about_me"] = about
    if add_price:
        data["additional_price"] = add_price
    if inferred_zone:
        data["filter_city"] = inferred_zone
    if min_str:
        data["min_spend"] = min_str
    if main_product:
        data["main_product"] = main_product[:200]
    return data


def _missing_required_columns(header_to_col: dict[str, int]) -> list[str]:
    missing: list[str] = []
    if "name" not in header_to_col:
        missing.append("name / 企业名称")
    if "lat" not in header_to_col:
        missing.append("lat / 纬度")
    if "lng" not in header_to_col:
        missing.append("lng / 经度")
    if "phone" not in header_to_col:
        missing.append("phone / 联系电话")
    has_address = "address" in header_to_col
    has_location_parts = any(
        header_to_col.get(f) is not None for f in ("province", "city", "district")
    )
    if not has_address and not has_location_parts:
        missing.append("address / 详细地址 (or 省份 + 城市 + 区县)")
    return missing


def parse_shop_import_excel(raw_bytes: bytes) -> tuple[list[tuple[int, dict[str, Any]]], str | None]:
    """
    Returns (list of (excel_row_number, payload_dict)), error_message_or_none.
    Rows with empty name are skipped. Stops after MAX_IMPORT_ROWS data rows.
    """
    try:
        wb = load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as e:
        return [], f"Invalid Excel file: {e}"

    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header_row = next(it, None)
        if not header_row:
            return [], "Empty spreadsheet"

        header_to_col: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            h = _norm_header(cell)
            if not h:
                continue
            field = _HEADER_TO_FIELD.get(h)
            if field and field not in header_to_col:
                header_to_col[field] = idx

        missing = _missing_required_columns(header_to_col)
        if missing:
            return [], "Missing required columns: " + ", ".join(missing)

        out: list[tuple[int, dict[str, Any]]] = []
        for row_idx, row in enumerate(it, start=2):
            if len(out) >= MAX_IMPORT_ROWS:
                break
            if not row:
                continue
            payload = _row_to_payload(header_to_col, row, header_row)
            name = (payload.get("name") or "").strip()
            if not name:
                continue
            out.append((row_idx, payload))
        return out, None
    finally:
        wb.close()


def build_template_workbook_bytes() -> bytes:
    """Template with English headers + a second reference row for Chinese industrial-belt sheets."""
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("factories")
    else:
        ws.title = "factories"
    headers = [
        "name",
        "address",
        "phone",
        "lat",
        "lng",
        "main_product",
        "filter_city",
        "about_me",
        "badge_text",
        "min_spend",
    ]
    ws.append(headers)
    ws.append(
        [
            "企业名称→name",
            "详细地址→address",
            "联系电话→phone",
            "纬度→lat",
            "经度→lng",
            "主营产品→main_product",
            "Pearl River Delta",
            "简介; 信用代码/注册资本等自动写入",
            "",
            "1-4 MOQ tier",
        ]
    )
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
